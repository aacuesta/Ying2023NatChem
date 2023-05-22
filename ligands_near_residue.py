#!/usr/bin/python
import sys, math
import openpyxl as op
import ftplib
import os, subprocess
import gzip
import time
import datetime
import platform


#program will take three things as arguments: The location of the PDB files, a list of names of the pdbfiles,
#and the path/filename for the output
#This version (v3_3) aims to use a less restrictive set of ligands and will not use Xiaobo's set of bad ligands
#It will only use those ligands which are explicitly indicated in the script. It does not remove nucleotides explicitly
#Version 3_3 does not remove phosphate containing ligands (unlike V3_2).


#pdbpath = sys.argv[1]
#pdblist= sys.argv[2]
#outputname = sys.argv[3]
#badligpath = sys.argv[4]

def dist (a1,a2):
  return math.sqrt((a1[0]-a2[0])*(a1[0]-a2[0])+(a1[1]-a2[1])*(a1[1]-a2[1])+(a1[2]-a2[2])*(a1[2]-a2[2]))


def get_pdb_file(pdbid, targetdir = ""):
    start_time = time.time()
    pdb = pdbid[0:4]
    #basepath = "/Users/acuesta/Documents/PDB/"

    target = targetdir + "/" + pdb + ".pdb"
    if os.path.exists(target):
        #print("%s exists!" % target)
        return target
    else:
        print("%s not found. Downloading..." % pdb)
        try:
            ftp = ftplib.FTP("ftp.rcsb.org") #pub/pdb
            ftp.login()
            ftp.cwd("/pub/pdb/data/structures/all/pdb")

            gzpdb = open("%spdb%s.pdb.gz" % (targetdir, pdb), 'wb')
            response = ftp.retrbinary('RETR pdb%s.ent.gz' % pdb.lower(), gzpdb.write)

            if not response.startswith('226-File successfully transferred'):
                print("%s download failed." % pdbid)
                print(response)
                gzpdb.close()
            else:
                print("%s downloaded." % pdbid)
                gzpdb.close()
                input = gzip.GzipFile("%spdb%s.pdb.gz" % (targetdir, pdb), 'rb')
                uncomp = input.read()
                input.close()

                with open(target, 'wb') as outpdbfile:
                    outpdbfile.write(uncomp)
                os.remove("%spdb%s.pdb.gz" % (targetdir, pdb))






        except ftplib.all_errors as e:
            print("FTP error:", e)
            raise


    #print("get_pdb_file() complete after %s seconds." % (time.time()-start_time))
    return target


#Returns a dictionary containing the chain and pdb residue numbers of interest.
def match_uniprot_pdb_num(up_res_num, uid, pdbid, query, chain_resnum, chain_resname, pthreshold, ethreshold):
    start_time = time.time()

    chain = pdbid[-1]

    #save the uniprotID of interest to file.
    #will feed this file to blastp later
    temp_uid_input = open("query.txt", 'w')
    temp_uid_input.write(query)
    temp_uid_input.close()

    #save the pdbid of interest to file.
    #must be in the format of PDBID_Chain
    #will feed this to blastp later
    temp_pdb_limit = open("subject.txt", 'w')
    #print(chain_resname[chain])
    try:
        temp_pdb_limit.write(chain_resname[chain])
        temp_pdb_limit.close()
    except KeyError:
        print("%s: Chain numbering mismatch between blast and pdb file. Skipping." % pdbid)
        temp_pdb_limit.close()
        raise



    command_line = ['blastp', '-query', "query.txt", '-subject', "subject.txt", '-outfmt', "10 qstart qend qseq sstart send sseq evalue pident"]
    com = subprocess.run(command_line, capture_output=True, text=True)
    results = com.stdout
    error = com.stderr
    #print(results)
    #print(error)

    output = []
    for line in results.split("\n"):
        output.append(line.split(","))

    chain_res = {}



    for line in output:
        #print(line)
        if len(line) > 1:
            querystart = int(line[0])
            queryend = int(line[1])

            queryseq = line[2]
            subjseq = line[5]

            subjstart = int(line[3])

            pos = querystart - 1 #It will increment to the correct starting residue number when the loop begins.
            stringpos = 0

            evalue = line[-2]
            if "-" in evalue:
                evalue = int(evalue[evalue.index("-") + 1 :])
            elif float(evalue) <= 1:
                evalue = float(ethreshold)
            else:
                evalue = float(evalue)


            percent = float(line[-1])

            #print(line)
            if queryend >= up_res_num and up_res_num >= querystart and evalue >= float(ethreshold) and percent >= float(pthreshold):
                for index, aa in enumerate(queryseq):
                    if aa != "-":
                        pos +=1
                    if pos == up_res_num:
                        if subjseq[index] != "-":

                            subj_index = subjstart - 1 + index - subjseq[:index].count("-") #"subjstart" is 1-indexed. Must subtract by one to make it 0-indexed, like "index"
                            #print(subj_index)
                            #print(chain)
                            subj_pos = chain_resnum[chain][subj_index]
                            chain_res[chain] = subj_pos
                            #print(subj_pos)
                            break


    #command_line = ['blastp', '-db', 'pdbaa', '-query', "query.txt", '-seqidlist', "pdblist.txt"]#, '-outfmt', "10 qstart qend qseq sstart send sseq sseqid"]
    #com = subprocess.run(command_line, capture_output=True, text=True)
    #results = com.stdout
    #error = com.stderr
    #print(results)

    #print("Blast searching done after %s seconds." % (time.time()-start_time))
    return chain_res

"""
Reads in PDB file
Returns 5 dictionary: aa position, aa abbreviation, ligand position, ligand name, chain-uniprot
"""
def parse_pdb(pdbfilename):
    start_time = time.time()

    """Open the file so that it can be accessed"""
    f = open(pdbfilename)
    pdbfile = f.readlines()
    f.close

    """dictionaries to store info from the pdb file"""
    res_position = {}
    res_abr ={}

    lig_position = {}
    lig_name = {}

    chain_dbid = {}

    chain_sequence = {}
    chain_res_num = {}

    header_string = ""
    chain_header = {}

    AA_single = {'ALA' : 'A', 'CYS': 'C', 'ASP': 'D', 'GLU' : 'E', 'PHE': 'F', 'GLY' : 'G',
        'HIS' : 'H', 'ILE' : 'I', 'LYS' : 'K', 'LEU' :'L', 'MET':'M', 'ASN':'N', 'PRO':'P',
        'GLN':'Q', 'ARG':'R', 'SER':'S', 'THR':'T', 'VAL':'V', 'TRP':'W', 'TYR':'Y', 'UNK' : 'X',
        'ASX': 'B', 'GLX':'Z', 'SEC': 'U', 'CSE':'U', 'PYL':'O'}


    for line in pdbfile:
        #print(line)
        if line[0:4] == "ATOM" or line[0:6] == "HETATM":
            abrv = line[17:20].strip().upper()
            chain = line[21]
            res_num = int(line[22:26])
            atom = line[12:16].strip().upper()
            position = (float(line[30:38]),float(line[38:46]),float(line[46:54]))

            if line[0:4] == "ATOM":
                try:
                    res_position[chain][res_num][atom] = position

                except KeyError:
                    try:
                        res_position[chain][res_num] = {atom : position}
                        res_abr[chain][res_num] = abrv
                        if res_num not in chain_res_num[chain]:
                            chain_res_num[chain].append(res_num)
                            chain_sequence[chain] += AA_single[abrv]
                    except KeyError:
                        try:
                            chain_sequence[chain] = AA_single[abrv]
                        except KeyError as k:
                            #print('Not an amino acid. Skip.')
                            #print(k)
                            continue
                        chain_res_num[chain] = [res_num]
                        res_position[chain] = {res_num : {atom : position}}
                        res_abr[chain] = {res_num : abrv}

            else:
                try:
                    lig_position[chain][abrv][atom] = position
                except KeyError:
                    try:
                        lig_position[chain][abrv] = {atom : position}
                    except KeyError:
                        lig_position[chain] = {abrv : {atom : position}}

            #print(res_position)

        elif line[0:6] == "DBREF ":
            #Contains related DB information
            #PDB Chain ID, DB Abbr, DB ID, Entry Description
            chain = line[12]
            db_abbr = line[26:32].strip()
            db_id = line[33:41].strip()
            desc = line[42:54].strip()

            if db_abbr == 'UNP':
                chain_dbid[chain] = {"UNP" : db_id}
            elif db_abbr == "GB":
                chain_dbid[chain] = {"GB" : db_id}

        elif line[0:6] == "HETNAM":
            id = line[11:14].strip()
            name = line[15:].strip()
            if id not in lig_name.keys():
                lig_name[id] = name
            else:
                lig_name[id] += name

        elif ((line[0:6] == "COMPND") and (line[11:20] == "MOLECULE:")):
            header_string += line[11:len(line)].strip()
        elif ((line[0:6] == "COMPND") and (line[11:17] == "CHAIN:")):
            header_string += line[11:len(line)].strip()

    #print(header_string)
    headerlist = header_string.split("MOLECULE:")
    for each in headerlist[1:]:
        #print(each)
        #print(headerlist)
        header_segments = each.split("CHAIN:")
        #print(header_segments)
        chain_string = header_segments[1].strip()
        molecule_name = header_segments[0].strip()
        chain_header[chain_string] = molecule_name

    #print("PDB file parsed after %s seconds." % (time.time()-start_time))
    return res_position, res_abr, lig_position, lig_name, chain_dbid, chain_res_num, chain_sequence, chain_header

#Output: {ResName : {ResNum :{Chain {Lig Abbrv : (Lig Name, Distance)}}}}
def aa_ligand_dist(aa_pos, aa_name, lig_pos, lig_name, chain_res_num, dist_threshold = 10):
    start_time = time.time()

    output_dict = {} #{ResName : {ResNum :{Chain {Lig Abbrv : (Lig Name, Distance)}}}}

    allowed_residues = {"CYS" : ["SG"], "ASP" : ["OD1", "OD2"], "GLU": ["OE1", "OE2"],
        "HIS" : ["ND1", "NE2"], "LYS" : ["NZ"], "ASN" : ["OD1", "ND2"], "GLN" : ["OE1", "NE2"],
        "ARG" : ["NH1", "NH2"], "SER" : ["OG"], "THR" : ["OG1"], "TYR" : ["OH"]}

    lig_dist = {}
    lig_neighbor = {}

    #If there is no ligand, return a placeholder and don't bother with the rest of this FUNCTION
    if lig_name == {}:
        #print("Found empty lig dict")
        for aa_chain in chain_res_num.keys():
            resnum = chain_res_num[aa_chain]
            distance = "N/A"
            lig = "N/A"

            try:
                aa_abrv = aa_name[aa_chain][resnum]
            except KeyError:
                #print("aa_chain is %s" % aa_chain)
                #print("resnum is %s" % resnum)
                #print(aa_name)
                print("Error. Residue number not found in PDB file. Check for possible numbering error with this file:")
                raise

            try:
                output_dict[aa_abrv][resnum][aa_chain][lig] = ("N/A", distance)
            except KeyError:
                try:
                    output_dict[aa_abrv][resnum][aa_chain] = {lig : ("N/A", distance)}
                except KeyError:
                    try:
                        output_dict[aa_abrv][resnum] = {aa_chain: {lig : ("N/A", distance)}}
                    except KeyError:
                        output_dict[aa_abrv] = {resnum : {aa_chain : {lig: ("N/A", distance)}}}
        return output_dict

    #If there are good ligands, proceed with the rest of the function
    for chain in lig_pos.keys():
        for lig in lig_pos[chain].keys():
            for aa_chain in chain_res_num.keys():
                resnum = chain_res_num[aa_chain]
                #print(aa_name)
                #print(aa_chain)
                #print(resnum)
                #print(aa_name.keys())
                #print(aa_name[aa_chain].keys())
                try:
                    aa_abrv = aa_name[aa_chain][resnum]
                except KeyError:
                    #print("aa_chain is %s" % aa_chain)
                    #print("resnum is %s" % resnum)
                    #print(aa_name)
                    print("Error. Residue number not found in PDB file. Check for possible numbering error with this file:")
                    raise

                for atom in aa_pos[aa_chain][resnum].keys():
                    aa_xyz = aa_pos[aa_chain][resnum][atom]

                    try: #Assume that the residue of interest is a nucleophilic residue
                        if atom in allowed_residues[aa_abrv]:

                            for ligatom in lig_pos[chain][lig]:
                                lig_xyz = lig_pos[chain][lig][ligatom]
                                distance = dist(aa_xyz, lig_xyz)

                                try:
                                    if lig_dist[lig] > distance:
                                        lig_dist[lig] = distance
                                        lig_neighbor[lig] = (aa_abrv, resnum, aa_chain)
                                except:
                                    lig_dist[lig] = distance
                                    lig_neighbor[lig] = (aa_abrv, resnum, aa_chain)
                    except KeyError: #The residue of interest isnt nucleophilic. So just check all residue atoms then.
                        for ligatom in lig_pos[chain][lig]:
                            lig_xyz = lig_pos[chain][lig][ligatom]
                            distance = dist(aa_xyz, lig_xyz)

                            try:
                                if lig_dist[lig] > distance:
                                    lig_dist[lig] = distance
                                    lig_neighbor[lig] = (aa_abrv, resnum, aa_chain)
                            except:
                                lig_dist[lig] = distance
                                lig_neighbor[lig] = (aa_abrv, resnum, aa_chain)

    for lig in lig_dist.keys():
        resname = lig_neighbor[lig][0]
        resnum = lig_neighbor[lig][1]
        chain = lig_neighbor[lig][2]
        distance = lig_dist[lig]

        if distance <= dist_threshold:
            try:
                output_dict[resname][resnum][chain][lig] = (lig_name[lig], distance)
            except KeyError:
                try:
                    output_dict[resname][resnum][chain] = {lig : (lig_name[lig], distance)}
                except KeyError:
                    try:
                        output_dict[resname][resnum] = {chain: {lig : (lig_name[lig], distance)}}
                    except KeyError:
                        output_dict[resname] = {resnum : {chain : {lig: (lig_name[lig], distance)}}}

    #print("Distance calculations done after %s seconds." % (time.time()-start_time)    )
    return output_dict

"""
Accept a PDB file and return a list of unique useful ligands.
Useful ligands are those that are not salts, crystalographic buffers or non-competable
cofactors like heme, and entries that are part of the chain that are listed as heteroatoms
such as non-standard amino acids or amino acids with post-translational modifications.

Items filtered include:
MSE: Selenomethionine, PTR: Phosphotyrosine, ORN: Ornithine, 4BF: 4-Bromo-Phenylalanine,
TYS: Sulfotyrosine, TPO: Phosphothreonine, SEP: Phosphoserine, CCS: CarboxymethylCysteine,
CSW: Cysteine Dioxide, HEM: Heme, HOH: Water, CSX: S Oxy Cysteine, HDD: CIS-HEME D HYDROXYCHLORIN GAMMA-SPIROLACTONE
SRM: Siroheme


Ions such as:
SO4: Sulfate, CA: Calcium, CL:Chloride, NA: Sodium, PO4: Phosphate, MN, Manganese, K: Potassium,
FE: Iron,  F: Fluoride, LI: Lithium, MN3: Manganese 3+, NO2: Nitrate, PD: Palladium ion, VO4: Vanadate
YT3: Yttrium(III), W: Tungsten, WO4, Tungstate, 3PO: Triphosphate, MGF: Trifluoromagnesate
SO3: Sulfite, RB: Rubidium, POP: Pyrophosphate, PR: Praseodymium, PO3: Phosphite, OS: Osmium
MF4: Magnesium Tetraflouride, LU: Lutetium, PB: Lead(II), GD: Gadolinium, EMC: Ethylmercury
HG2: Dibromomercury, CS: Cesium, ALF: Tetrafluoroaluminate, AF3: Aluminum Flroride, PPK: (DIPHOSPHONO)AMINOPHOSPHONIC ACID
TMO: Trimethylamine Oxide, SAT: Sulfoacetic Acid, SIN: Succinic Acid, RH3: Rhodium (III), PI: Hydrogenphosphate
HO: Holmium, 6BP: Hexabromoplatinate, DVT: Decavanadate, ARS: Arsenic, SM: Samarium

PTM:
BMA: Beta-D-Mannose, CRQ: Fluorescent Protein Chromophore, MLZ: N-Methyl Lysine, MME: N Methyl Methionine
NAG: N-Acetyl-Glucosamine, SMC: S-MethylCysteine, OFM: Fluorescent Protein Chromophore,
OIM: Fluorescent Protein Chromophore, NRQ: Fluorescent Protein Chromophore, CRO: Fluorescent Protein Chromophore,
CH6: Chromophore, EYG: Chromophore, CGU: Gamma Carboxy glutamate

Crystallographic Artifacts and Detergents:
BOG: Beta-Octyl Glucoside, GOL: Glycerol, CIT: Citrate, CPS: CHAPS Detergent, DMF: Dimethylformamide,
DOD: Deuterated Water, FLC: Citrate, GAI: Guanidinium, NME: Methylamine, PCA: Pyroglutamic Acid (N terminal cyclized glutamine)
PEO: Hydrogen Peroxide, LMU: DODECYL-ALPHA-D-MALTOSIDE, MES: 2-(N-MORPHOLINO)-ETHANESULFONIC ACID, OXL: Oxalate
URE: Urea, TAM: Tris, BET: Trimethyl Glycine, TFA: Trifluoroacetic Acid, TGL: Triacylglycerol,
STE: Stearic Acid, TTN: Tartronate, PGO: Propanediol, 12P: Polyethlene glycol
P4C: Polyethylene, P3A: Phosphatidyl Glycerol, 7PH: Phosphatidic Acid, PLM: Palmitic Acid, OLA: Oleic Acid
DR6: Big big detergent, 37X: Octaglucose neopentyl glycol, HSJ: OCTYL BETA-L-TALOPYRANOSIDE, IDS: O2-SULFO-GLUCURONIC ACID
2PE: Nonaethylene glycol, OCT: N-Octane, MYR: Myristic Acid, SGM: Monothioglycerol, MOH: Methanol
DAO: Lauric Acid, GOA: Glycolic Acid, HEZ: Hexane-1,6-Diol, PEE: DOPE, LMT: Dodecyl-Beta-D-maltoside, SDS: Sodium Dodecul Sulfate
PLC: Diundecyl Phosphatidyl Choline, CDL: Cardiolipin , DMA: Dimethylallyl diphosphate, DKA: Decanoic Acid
ACM: Acetamide, PC1; Big lipid, PGA: Phosphoglycolic acid, FGL: 2-AMINOPROPANEDIOIC ACID, 1PG: PEG
COM: 1-THIOETHANESULFONIC ACID, OLC: (2R)-2,3-dihydroxypropyl (9Z)-octadec-9-enoate, DIO: Dioxin
BU2: 1,3-Butanediol, 52N: 1,2-dioctanoyl phosphatidyl epi-inositol (3,4)-bisphosphate,
PIE: 1,2-DIACYL-SN-GLYCERO-3-PHOSPHOINOSITOL, L2O: (2S,3R)-3-amino-2-hydroxy-5-methylhexanoic acid,
LYK: (2S)-2,6-diaminohexane-1,1-diol, DTU: (2R,3S)-1,4-DIMERCAPTOBUTANE-2,3-DIOL, TMA: Tetramethylammonium,
PPI: Propanoic Acid, 7PE: PEG Fragment, PE3: PEG, PPF: PHOSPHONOFORMIC ACID, PGH: PHOSPHOGLYCOLOHYDROXAMIC ACID
IPH: Phenol, C15: N-DODECYL-N,N-DIMETHYL-3-AMMONIO-1-PROPANESULFONATE, HEX: Hexane, 6NA: Hexanoic Acid
P33: PEG330, B7G: HEPTYL-BETA-D-GLUCOPYRANOSIDE (B7G), XPE: Decaethylene Glycol, ETE: Some Polyester
B3P: DiTris, 2PG: 2-PHOSPHOGLYCERIC ACID, MXE: 2-MethoxyEthanol, ETX: 2-EthoxyEthanol, 3PP: 3-PHOSPHONOPROPANOIC ACID
PG5: Some polyester, P4G: Some Polyester, BU1: 1,4-Butanediol, DXE: Dimethoxyethane, BU3: 2,3-Butanediol,
0V5: (2R)-2-(phosphonooxy)propanoic acid, TBU: TERTIARY-BUTYL ALCOHOL, PLD: DIUNDECYL PHOSPHATIDYL CHOLINE
PEF: Big big lipid, EPH: Big lipid, OPE: Cholamine Phosphoric Acid, 15P: PEG


Buffers:
MPO: 3 N-Morpholino Propane Sulfonic Acid, CPQ: Deoxy-Bigchap, NHE: CHES

Sugars:
SUC: Sucrose, SPV: Sulfopyruvate, PYR: Pyruvic Acid, PEP: Phosphoenolpyruvate, SIA: Sialic Acid
MAL: Maltose, MLI: Malonate, MLT: Malate, PEQ: L-phospholactate, LMR: L-malate, IDR: Iduronic Acid
ICT: Isocitric Acid, ISD: Isoascorbic acid, 4IP: INOSITOL-(1,3,4,5)-TETRAKISPHOSPHATE,
F6P: Fructose-6-Phosphate, F6R: Fructose-6-Phosphate, MLA: Malonic Acid, TAR: Tartaric Acid
LAT: Beta-lactose, M6D: Mannose-6-Phosphate, BG6: Glucose-6-Phosphate, BGC: Beta-d-glucose
GLS: BETA-D-GLUCOPYRANOSE SPIROHYDANTOIN, GAL: Beta-d-galactose, MLR: Maltriose,
RAM: Rhamnose, FUC: Fucose, MAK: Ketomalonic Acid, MAN: Mannose, G6P: Alpha Glucose 6 Phosphate,
VG1: Glucose-6-P-1-Vanadate, G1P: Glucose-1-P, G16: Glucose-1,6-Bisphosphate, GLC: Glucose
TRE: Trehalose, GLG: ALPHA-D-GLUCOPYRANOSYL-2-CARBOXYLIC ACID AMIDE, GLA: Alpha galactose
FUL: 6-Deoxy-Beta-L-Galactose, G6D: 6-Deoxy-Alpha-D-Glucode, MAH: 3-HYDROXY-3-METHYL-GLUTARIC ACID
AKG: 2-oxoglutaric acid, 5MM: DiphosphonoMannitol, ALS: 2-AMINO-3-OXO-4-SULFO-BUTYRIC ACID,
FGP: 2-AMINO-3-HYDROXY-3-PHOSPHONOOXY-PROPIONIC ACID, NDG: 2-(ACETYLAMINO)-2-DEOXY-A-D-GLUCOPYRANOSE
NBG: 1-N-ACETYL-BETA-D-GLUCOSAMINE, 2FP: Linear FructoseDiphosphate, MA1: Dithioglucopyranose
13P: Dihydroxyacetonephosphate, SRT: S,R MESO-TARTARIC ACID, AGL: 4,6-DIDEOXY-4-AMINO-ALPHA-D-GLUCOSE
OAA: Oxaloacetate, IHP: INOSITOL HEXAKISPHOSPHATE, INS: Myo-Inositol, FUM: FUMARIC ACID,
IP5: Myoinositol pentaphosphate, I3P: Inositol triphosphate, I7P: Inositol Phosphate, I8P: Inositol Phosphate
GLF: Fluoro-Alpha-Deoxy-Glucose, M2P: Manitol-Diphosphate, ASO: Anhydrosorbitol

Biological Lipids:
FPP: Farnesyl Diphosphate, GRG: GeranylGeranyl Diphosphate, FPS: Farnesyl thiopyrophosphate
FAR: Farnesyl, EIC: Linoleic Acid, MGM: 3-AZAGeranylgeranyl diphosphate, FII: FPP ANALOG, PCW: Phospholipid
P3A: Phospholipid, P6L: Phospholipid, PEF: Phopholipid

Other:
UNX: Unknown, UNL: Unknown Ligand, CPT: Cisplatin, PUT: 1,4-DIAMINOBUTANE, ICF: Isofluorane, UNK: Unknown
"""
def getGoodLigands(pos_dict, name_dict):

    badligands = ['NBN', 'ENC', 'MNC', 'NPN', 'GSH', '0HH', 'GTX', 'BEN', 'OAR', 'PYX', 'CFC', 'ETF', 'TDG', 'ADC', 'RIP', 'CFO', 'OFO', '9DI', 'HPA', 'NOS', 'MET', 'LEU', 'DME', 'TRA', '1DA', 'O', 'MBG', 'SCR', 'AYA', 'ETA', 'HAE', 'MIC', 'DAS', 'BE2', 'DCE', 'ASP', 'HEA', 'LDA', 'SNN', 'IAS', 'HIC', 'BHD', 'EDR', 'SGN', 'NTP', 'PYJ', 'SAC', 'TSU', 'CYH', 'AOP', 'CRA', 'MBR', 'CTO', 'UMG', 'GUM', 'BLV', 'FMS', 'A2G', 'DAR', 'PMB', 'MMC', '2FU', 'PBP', 'INN', 'BRC', 'SSB', 'SNC', 'HGB', 'AJ3', 'ADN', 'BAM', 'NVA', 'H2S', 'ADE', 'ACA', 'MTA', 'DEC', 'NON', 'COH', 'HMC', 'TRI', 'PHG', 'CAD', 'PG6', 'SAH', 'NH3', 'DDQ', 'ABH', 'SO2', 'SET', 'NLE', 'DSS', 'BUQ', 'CYS', 'AR7', '0QE', 'PYZ', 'ITU', 'LA', 'V7O', 'PHQ', 'HRG', 'CMT', 'GER', 'CSS', 'LPB', 'BOM', 'ASG', 'GLU', '4MZ', '1MZ', 'HAR', 'DHM', 'UNA', 'GUP', 'DDZ', 'BO3', 'AAE', 'DTD', 'PBR', 'HLT', 'ALA', 'PRO', 'GCS', 'IPU', 'IUM', 'CDU', 'HEV', 'CRS', 'RCO', 'IOL', 'HDS', 'ASJ', 'CNN', 'MGS', 'MGU', 'DCY', 'ESA', 'CE', 'GCU', 'HXA', 'LAX', 'AZE', 'AIB', 'SE4', 'ATH', 'MOS', 'FRU', 'FGA', 'DAM', 'PUR', 'HNI', 'ARG', 'TL', 'GL2', 'GL5', 'GL7', 'GL9', 'CR1', 'CR6', 'SGA', 'MAG', 'ACN', '3OL', 'AEM', 'EMM', '3OM', 'IUR', 'IDK', 'URA', 'GLY', 'NTA', 'CDG', 'GM2', 'D12', 'TWT', 'AOA', 'OGA', 'BEZ', 'G2F', 'DO3', 'IL0', 'CXL', 'AOE', 'ANL', 'NA1', 'NAA', 'TBR', 'PFB', 'BRB', 'FEO', 'ELA', 'NNH', 'PP9', 'UNK', 'PEF', 'P6L', 'P3A', 'PCW', 'NHE', '15P', 'CPQ', 'SM', 'MPO', 'OPE', 'EPH', 'PEF', 'PLD', 'CGU', 'TMA', 'TBU', '0V5', 'BU3', 'DXE', 'ASO', 'M2P', 'BU1', 'GLF', 'P4G', 'PG5', 'I7P', 'I8P', '3PP', 'ETX', 'MXE', '2PG', 'B3P', 'ETE', 'ARS', 'EYG', 'I3P', 'IP5', 'XPE', 'DVT', 'FII', 'FUM', 'B7G', 'P33', 'HO', '6BP', 'HEX', '6NA', 'PI', 'INS', 'IHP', 'C15', 'OAA', 'IPH', 'PGH', 'PPF', '7PE', 'PE3', 'PPI', 'PUT', 'AGL', 'RH3', 'SRT', 'SIN', 'SAT', 'TMO', 'DTU', 'LYK', 'L2O', 'PPK', 'PIE', '52N', 'BU2', '13P', 'DIO', 'MA1', '2FP', 'NBG', 'OLC', 'COM', '1PG', 'NDG', 'ALS', 'FGP', 'FGL', '5MM', 'AKG', 'PGA', 'MGM', 'MAH', 'PC1', 'FUL', 'G6D', 'EIC', 'ACM', 'GLA', 'GLG', 'TRE', 'GLC', 'VG1', 'G1P', 'G16', 'G6P', 'MAK', 'MAN', 'RAM', 'FUC', 'AF3', 'MLR', 'GLS', 'GAL', 'BGC', 'M6D', 'BG6', 'LAT', 'CS', 'CH6', 'TAR', 'DKA', 'HG2', 'CPT', 'MLA', 'DMA', 'CDL', 'LMT', 'SDS', 'PLC', 'PEE', 'EMC', 'FAR', 'FPP', 'GRG', 'F6P', 'F6R', 'GD', 'HEZ', 'GOA', '4IP', 'ISD', 'ICT', 'IDR', 'LMR', 'DAO', 'PEQ', 'PB','LU', 'MF4', 'MLT', 'MLI', 'MAL', 'MOH', 'SGM', 'MYR', 'OCT', '2PE', 'IDU', 'SIA', 'IDS', '37X', 'HSJ', 'DR6', 'OLA', 'OS', 'PLM', '7PH', 'PO3', 'PEP', '12P', 'P4C', 'PR', 'POP', 'PYR', 'RB', 'PGO', 'TTN', 'STE', 'SPV', 'SO3', 'SUC', 'ALF', 'TGL', 'TFA', 'MGF', 'BET', 'TAM', '3PO', 'WO4', 'W', 'UNL', 'URE', 'OXL', 'YT3', 'VO4', 'MES', 'LMU', 'CRO', 'NRQ', 'UNX', 'OIM', 'OFM', 'PD', 'SMC', 'PEO', 'PCA', 'NO2', 'NME', 'NAG', 'MN3', 'MME', 'MLZ', 'LI', 'HDD', 'GAI', 'FLC', 'DOD', 'DMF', 'BMA', 'CRQ', 'BOG', 'CIT', 'CPS', 'CSX', 'F', '202', 'SO4', 'GOL', 'ZN', 'MG', 'CA', 'CL', 'NA', 'EDO', 'PO4', 'HEM', 'ACT', 'MN', 'K', 'PEG', 'FE', 'ACE', 'CU', 'MPD', 'NI', 'NH2', 'DMS', 'BME', 'TRS', 'CD', 'PG4', 'FMT', 'ACY', 'SEP', 'EPE', 'SF4', 'PGE', 'CO', 'TPO', 'IOD', 'FE2', 'FES', 'IMD', 'NO3', 'HG', 'CSO', 'PTR', 'IPA', '1PE', 'KCX', 'MRD', 'HEC', 'CMO', 'BR', 'CO3', 'CME', 'TLA', 'SCN', 'OXY', 'P6G', 'NCO', 'CAC', 'TYS', 'CSD', 'MLY', 'DAL', 'F3S', 'EOH', 'CU1', 'AZI', 'HED', 'M3L', 'DTT', 'NH4', 'HYP', 'OCS', 'CYN', 'FME', 'CAS', 'ALY', 'BTB', 'ABA', 'XE', 'CSW', 'BCT', 'SR', 'BA', 'NO', 'C8E', 'BCL', 'BEF', 'PE4', 'AU', 'CP', 'PT', 'YB', 'MSE', 'ORN', '4BF', 'CCS', 'HOH', 'OH']


    filtered_pos_dict = {}
    filtered_name_dict = {}

    for chain in pos_dict.keys():
        for lig in pos_dict[chain].keys():

            if lig not in badligands:
                filtered_name_dict[lig] = name_dict[lig]

                for atom in pos_dict[chain][lig].keys():
                    try:
                        filtered_pos_dict[chain][lig][atom] = pos_dict[chain][lig][atom]
                    except:
                        filtered_pos_dict[chain] = {lig : {atom : pos_dict[chain][lig][atom] }}


    return filtered_pos_dict, filtered_name_dict

def removeLigandEntries(outputList):
    AdenineLigs = ['112', 'ATP', 'ADP', 'ANP', 'IMP', 'CMP', 'CXR', 'AGS', 'AVU', 'APR', 'ACP', 'A3P', 'AMP', 'PPS', '3AM']
    GuanineLigs = ['GSP', 'GNP', 'GDP', 'GCP', 'GP2', 'GTG', 'GTP', '5GP', 'G3D', 'G2R', 'G2Q', 'G1R', 'PCG', 'CGR', 'DGP']
    NADLigs = ['NAP', 'NDP', 'DN4', 'NAD', 'NMN', '2NF', 'NAI']
    FADLigs = ['FAD', 'FMN']
    UridineLigs = ['UPG', 'U5P','U2F', 'UDP', 'U2P', 'GDU', 'OMP', 'UMP','S5P', 'BMP', '6CN', 'NUP', '6AU', 'XMP', '5FU', 'CNU', '5BU', 'DUR', 'UD1']
    ThymidineLigs = ['TMP', 'TTP', 'TDP', 'ATM', '2DT']
    CytosineLigs = ['TKW', 'DCP']

    adenineOutput = []
    guanineOutput = []
    nadOutput = []
    fadOutput = []
    uridineOutput = []
    thymidineOutput = []
    cytosineOutput = []
    remainingOutput = []
    for line in outputList:
        if line[3] in AdenineLigs:
            adenineOutput.append(line)
        elif line[3] in GuanineLigs:
            guanineOutput.append(line)
        elif line[3] in NADLigs:
            nadOutput.append(line)
        elif line[3] in FADLigs:
            fadOutput.append(line)
        elif line[3] in UridineLigs:
            uridineOutput.append(line)
        elif line[3] in ThymidineLigs:
            thymidineOutput.append(line)
        elif line[3] in CytosineLigs:
            cytosineOutput.append(line)
        else:
            remainingOutput.append(line)

    return adenineOutput, guanineOutput, nadOutput, fadOutput, uridineOutput, thymidineOutput, cytosineOutput, remainingOutput

def write_output(out_dict, KR_dict, complex_dict, eval_dict, percent_dict, info_dict, filepath, outputpath, settings):
    inbook = op.load_workbook(filepath)
    outbook = op.Workbook()

    insheet = inbook.active
    outsheet = outbook.active


    OS = platform.system()
    #print (OS)
    #if OS == 'Darwin':
    separator = '/'
    #elif OS =='Windows':
        #separator = '\\'
    #else:
        #pass

    filename = filepath.split(separator)[-1]
    #print(separator)
    #print(filename)
    header = []
    for cell in insheet[1]:
        header.append(cell.value)


    uid_index = header.index("Protein")
    pos_index = header.index("Position")

    new_header = header + ["PDB ID", "PDB Protein Name", "PDB Residue", "PDB Chain",  "PDB Position", "Ligand ID", "Ligand Name",  "Distance", "E-value", "Percent Identity"]
    new_header += ["Neighboring Lys/Arg?", "Lys/Arg Distance"]
    new_header += ["Protein-Protein Interface", "Partner UniprotID", "Partner Name"]

    for index, item in enumerate(new_header):
        outsheet.cell(row=1, column=index+1).value=item

    for index, row in enumerate(insheet):
        #print(len(row))
        if row != insheet[1]:

            data = []
            for cell in row:
                data.append(cell.value)

            uniprotid = data[uid_index]
            data[uid_index] = "=Hyperlink(\"https://www.uniprot.org/uniprot/%s\", \"%s\")" % (uniprotid, uniprotid)
            try:
                uid = row[uid_index].value[0:6]
            except TypeError: #this error occurs when "None" is present in uid cell
                continue #If you find "None" in the uid cell, skip everything below, and move on to the next row.
            pos = int(row[pos_index].value)
            if uid in out_dict.keys():
                if pos in out_dict[uid].keys():
                    for pdbid in out_dict[uid][pos].keys():
                        for pdb_res_name in out_dict[uid][pos][pdbid].keys():
                            for pdb_res_num in out_dict[uid][pos][pdbid][pdb_res_name].keys():
                                for chain in out_dict[uid][pos][pdbid][pdb_res_name][pdb_res_num].keys():
                                    for lig_abrv in out_dict[uid][pos][pdbid][pdb_res_name][pdb_res_num][chain].keys():

                                        lig_name = out_dict[uid][pos][pdbid][pdb_res_name][pdb_res_num][chain][lig_abrv][0]
                                        distance = out_dict[uid][pos][pdbid][pdb_res_name][pdb_res_num][chain][lig_abrv][1]

                                        #Prep K/R neighbors for output
                                        KRs = KR_dict[uid][pos][pdbid]
                                        KRout = ""
                                        KRdistout = ""
                                        if len(KRs.keys()) == 0:
                                            KRout = "No;"
                                            KRdistout = "N/A;"
                                        else:
                                            try:
                                                for partner_chain in KRs[pdb_res_name][pdb_res_num].keys():
                                                    for KR in KRs[pdb_res_name][pdb_res_num][partner_chain].keys():
                                                        for KRpos in KRs[pdb_res_name][pdb_res_num][partner_chain][KR].keys():
                                                            KRdist = KRs[pdb_res_name][pdb_res_num][partner_chain][KR][KRpos]
                                                            KRout += "%s %s %s;" % (KR, partner_chain, KRpos)
                                                            KRdistout += "%s;" % (KRdist)
                                            except:
                                                print(pdbid)
                                                print(KRs)
                                                print(KRs[pdb_res_name][pdb_res_num].keys())
                                        KRout = KRout[:-1]
                                        KRdistout = KRdistout[:-1]

                                        ppi = complex_dict[uid][pos][pdbid]

                                        ppitype = ""
                                        partnerout = ""
                                        partnername = ""
                                        if len(ppi.keys()) == 0:
                                            ppitype = "No"
                                            partnerout = "N/A"
                                            partnername = "N/A"
                                        else:
                                            for ppitype in ppi[pdb_res_name][pdb_res_num][chain].keys():
                                                for partner_chain in ppi[pdb_res_name][pdb_res_num][chain][ppitype].keys():

                                                    partner_ids = ppi[pdb_res_name][pdb_res_num][chain][ppitype][partner_chain]

                                                    ppiout = type


                                                    for each in partner_ids:
                                                        try:
                                                            partnerout += each + ";"
                                                            #print(info_dict[uid][pos][pdbid])
                                                            for chainset in info_dict[uid][pos][pdbid].keys():
                                                                if partner_chain in chainset:
                                                                    partnername += info_dict[uid][pos][pdbid][chainset] + ";"
                                                        except KeyError as e:
                                                            #print(e)
                                                            pass


                                        for chains in info_dict[uid][pos][pdbid].keys():
                                            if chain in chains:
                                                molname = info_dict[uid][pos][pdbid][chains]
                                            #print(info_dict)
                                            #print(chains)
                                        try:
                                            evalue = eval_dict[uid][pos][pdbid]
                                        except:
                                            evalue = "Not found"
                                        try:
                                            percentID = percent_dict[uid][pos][pdbid]
                                        except:
                                            percentID = "Not found"
                                        pdblink = "=Hyperlink(\"https://rcsb.org/structure/%s\", \"%s\")" % (pdbid[0:4], pdbid[0:4])
                                        new_results = [pdblink, molname, pdb_res_name, chain, pdb_res_num, lig_abrv, lig_name,  distance, evalue, percentID, KRout, KRdistout, ppitype, partnerout, partnername]
                                        #print(new_results)
                                        outsheet.append(data + new_results)


    if ".xlsx" in filename:
        outbook.save(filename = outputpath + filename[:-5]+ "with ligands.xlsx")
    else:
        outbook.save(filename = filename + "with ligands.xlsx")
    #if status != "":
    #    status("Done!")

#Get a dictionary with all residues with any atom within X angstroms of any atom in a query residue of interest
def find_neighbors(aa_pos, aa_abrv, qdict, dist_threshold, pdbid_chain):
    query_chain = pdbid_chain[-1]
    query_xyzs = aa_pos[query_chain][qdict[query_chain]]

    neighbors = {}

    for chain in aa_pos.keys():
        for pos in aa_pos[chain].keys():
            for atom in aa_pos[chain][pos].keys():

                for qatom in query_xyzs.keys():
                    neigh_xyz = aa_pos[chain][pos][atom]
                    query_xyz = query_xyzs[qatom]

                    distance = dist(neigh_xyz, query_xyz)

                    if distance <= dist_threshold:
                        try:
                            neighbors[chain][pos] = aa_pos[chain][pos]
                        except KeyError:
                            neighbors[chain] = {pos : aa_pos[chain][pos]}
    #print(pdbid_chain)
    #print(qdict[query_chain])
    #print(query_xyzs)

    #print(neighbors)


    return neighbors

def KRnearY(all, allabr, allnear, query, dist_threshold, pdbid_chain):
    query_chain = pdbid_chain[-1]
    query_num = query[query_chain]
    query_abr = allabr[query_chain][query_num]
    query_xyzs = all[query_chain][query_num]

    output = {}

    for chain in allnear.keys():
        for pos in allnear[chain].keys():
            for atom in allnear[chain][pos].keys():
                #print(allabr[chain][pos])
                neighborname = allabr[chain][pos]
                if (neighborname == "LYS" and atom == "NZ") or (neighborname == "ARG" and (atom == "NH1" or atom == "NH2")):

                    for qatom in query_xyzs.keys():
                        if qatom == "OH":

                            neigh_xyz = allnear[chain][pos][atom]
                            query_xyz = query_xyzs[qatom]

                            distance = round(dist(neigh_xyz, query_xyz),2)

                            if distance <= dist_threshold:
                                try:
                                    if output[query_abr][query_num][chain][neighborname][pos] > distance:
                                        output[query_abr][query_num][chain][neighborname][pos] = distance
                                except KeyError:
                                    try:
                                        output[query_abr][query_num][chain][neighborname] = {pos : distance}
                                    except KeyError:
                                        try:
                                            output[query_abr][query_num][chain] = {neighborname : {pos : distance}}
                                        except KeyError:
                                            try:
                                                output[query_abr][query_num] = {chain : {neighborname : {pos : distance}}}
                                            except KeyError:
                                                output[query_abr] = {query_num : {chain : {neighborname : {pos : distance}}}}

    #print(output)
    return output

def ppicheck(all, allabr, allnear, query, dist_threshold, pdbid_chain, refdb):
    #Check the chain of all neighboring residues
    #Is it the same chain as the query residue?
    #If yes, then it is not at a protein-protein interface
    #If no, then compare the uniprot IDs of the different chains
    #Use the name in the header of the PDB file

    query_chain = pdbid_chain[-1]
    query_num = query[query_chain]
    query_abr = allabr[query_chain][query_num]
    query_xyzs = all[query_chain][query_num]

    near_chains = allnear.keys()


    dimer_type = ""
    partner_id = ""

    output = {}

    #Not a protein interface
    #Return a placeholder
    if len(near_chains) == 1:
        return output
        #dimer_type = "No"
        #partner_id = [""]
        #output[query_abr] = {query_num : {query_chain : {dimer_type : partner_id}}}

    elif len(near_chains) > 1:
        #Multiple chains near the residue of interest, must be at an interface
        #Is it a homodimer or heterodimer
        for chain in near_chains:
            if chain != query_chain:
                if refdb[chain] == refdb[query_chain]:
                    dimer_type = "Homodimer"
                else:
                    dimer_type = "Heterodimer"
                try:
                    partner_id = refdb[chain]["UNP"]
                except KeyError:
                    for each in refdb.values():
                        partner_id += each + ";"

                try:
                    if partner_id not in output[query_abr][query_num][query_chain][dimer_type][chain]:
                        output[query_abr][query_num][query_chain][dimer_type][chain] += partner_id
                except KeyError:
                    try:
                        output[query_abr][query_num][query_chain] = {dimer_type : {chain : [partner_id]}}
                    except KeyError:
                        try:
                            output[query_abr][query_num] = {query_chain : {dimer_type : {chain : [partner_id]}}}
                        except KeyError:
                            output[query_abr]= {query_num :{query_chain : {dimer_type : {chain : [partner_id]}}}}


    #print(output)
    return output

def progress(progbar, status, count, total, start):
    if progbar != None:
        progbar(1 + count/total * 100)
        time_remaining = str(datetime.timedelta(seconds=(round((time.time() - start)/count * (total - count), 1)))).split(".")[0]
        status ("Performing ligand search. Estimated time remaining: %s" % time_remaining)


def find_ligands(uniprot_pdb_map, evalues_dict, percents_dict, query_seq_dict, inputpath, settings, outputpath, progbar = None, status = ''):
    start_time = time.time()
    output_dict = {}
    header_dict = {}
    KR_dict = {}
    complex_dict = {}
    failed_pdbs = []

    threshold = float(settings['distance_cutoff'])
    pdb_path = settings['pdb_path'] + "/"

    evalthreshold = float(settings["eval_cutoff"])
    percentthreshold = float(settings["min_percent"])

    total_pdbs = 0
    for uniprotid in uniprot_pdb_map.keys():
        for pos in uniprot_pdb_map[uniprotid].keys():
            for pdb in uniprot_pdb_map[uniprotid][pos]:
                total_pdbs += 1

    count = 0.
    for uniprotid in uniprot_pdb_map.keys():
        for pos in uniprot_pdb_map[uniprotid].keys():
            for pdb in uniprot_pdb_map[uniprotid][pos]:

                count +=1

                #Get all of the information out of the pdb file
                #Then figure out which residues I care about (match_uniprot_pdb_num)
                #If it is a nucleophilic residue, then get the shortest distance between the nucleophilic atom and each ligand atom
                ##Report the two atoms and the distance between them


                if pdb[0:4] not in failed_pdbs:
                    try:
                        pdb_filename = get_pdb_file(pdb, pdb_path)

                        #print(pdb_filename)
                        aa_pos, aa_abr, lig_pos, lig_name, chain_refdbs, chain_resnum_list, chain_abrv_list, chain_header = parse_pdb(pdb_filename)
                        filt_lig_pos, filt_lig_name = getGoodLigands(lig_pos, lig_name)

                        #print(chain_header)

                        #if filt_lig_name == {}:
                        #    progress(progbar, status, count, total_pdbs, start_time)
                        #    continue
                        #print(query_seq_dict)
                        query = query_seq_dict[uniprotid][pos]

                        pdb_numbering  = match_uniprot_pdb_num(pos, uniprotid, pdb, query, chain_resnum_list, chain_abrv_list, percentthreshold, evalthreshold)

                        results = aa_ligand_dist(aa_pos, aa_abr, filt_lig_pos, filt_lig_name, pdb_numbering, threshold)

                        neighbors = find_neighbors(aa_pos, aa_abr, pdb_numbering, 6, pdb)

                        #Only do this step if the query residue in this file is a tyrosine
                        #This step looks for neighboring lysines and arginines that can activate the tyrosine
                        if aa_abr[pdb[-1]][pdb_numbering[pdb[-1]]] == "TYR" :
                            KRneighbors = KRnearY(aa_pos, aa_abr, neighbors, pdb_numbering, 5, pdb)
                        else:
                            KRneighbors = {}

                        #Look for protein-protein interactions near the tyrosine in this step
                        ppis = ppicheck(aa_pos, aa_abr, neighbors, pdb_numbering, 6, pdb, chain_refdbs)

                        if results == {}:
                            progress(progbar, status, count, total_pdbs, start_time)
                            continue
                    except Exception as e:
                        print(e)
                        failed_pdbs.append(pdb[0:4])
                        progress(progbar, status, count, total_pdbs, start_time)
                        continue

                    try:
                        output_dict[uniprotid][pos][pdb] = results
                        header_dict[uniprotid][pos][pdb] = chain_header
                        KR_dict[uniprotid][pos][pdb] = KRneighbors
                        complex_dict[uniprotid][pos][pdb] = ppis
                        #print(output_dict)
                    except KeyError as e:
                        try:
                            output_dict[uniprotid][pos] = {pdb : results}
                            header_dict[uniprotid][pos] = {pdb : chain_header}
                            KR_dict[uniprotid][pos] = {pdb : KRneighbors}
                            complex_dict[uniprotid][pos] = {pdb : ppis}
                            #print(output_dict)
                        except:
                            output_dict[uniprotid]= {pos: {pdb : results}}
                            header_dict[uniprotid] = {pos : {pdb : chain_header}}
                            KR_dict[uniprotid] = {pos : {pdb : KRneighbors}}
                            complex_dict[uniprotid] = {pos : {pdb : ppis}}
                            #print(output_dict)
                    #print(output_dict)
                    progress(progbar, status, count, total_pdbs, start_time)




    write_output(output_dict, KR_dict, complex_dict, evalues_dict, percents_dict, header_dict, inputpath, outputpath, settings)
    status("Done!")
    print("Elapsed time: %s" % (time.time()-start_time))


"""
MAIN


ligandList = []
for file in pdbs:
    print strftime("%Y-%m-%d %H:%M:%S")
    pdbfilepath = pdbpath + "/" + file[0:4] + ".pdb"

    ligands, ligandNamesDict = getGoodLigands(pdbfilepath)
    print file
    print ligands

    for ligandName in ligands:

        header, lys, lig, refdb, ligSerialDict, connectDict = getResidues(pdbfilepath, ligandName)
        lys10 = kNearLigand(lys, lig)
        activatedLys = formatK(lys10)
        print activatedLys
        if len(activatedLys) > 0:
            outputLog = makeOutput(header, activatedLys, outputLog, ligandName, refdb, ligandNamesDict[ligandName])
        ligandList.append(ligandName)




adenineTable, guanineTable, nadTable, fadTable, uridineTable, thymidineTable, cytosineTable, remainingTable = removeLigandEntries(outputLog)
"""
