import sys, os, subprocess, csv
from openpyxl import load_workbook
from openpyxl import Workbook
import ligands_near_residue as lnr
import time
import threading
#inputpath = sys.argv[1]
#outputpath = sys.argv[2]

#open the file containing the uniprotID and resn
#make a dictionary containing uniprotID as key and list of resn as value

#get PDB ID of

if __name__ == '__main__':
	main()

def parse_input(inputname, progbar=None, status=""):
	output_dict = {}
	numlines = 0
	#print(inputname)
	suffix = inputname.strip()[inputname.index("."):].lower()
	if suffix == ".csv":
		with open(inputname, 'r') as inputfile:
			reader = csv.reader(inputfile)
			header = next(reader)
			uid_index = header.index("Protein")
			pos_index = header.index("Position")
			for row in reader:
				uid = row[uid_index][0:6]
				if uid != '':
					numlines +=1
					pos = int(row[pos_index])
					if uid not in output_dict.keys():
						output_dict[uid] = [pos]
					else:
						output_dict[uid].append(pos)

	elif suffix == ".xlsx" or suffix ==".xls":
		wb = load_workbook(inputname)
		ws = wb.active
		header = []
		for cell in ws[1]:
			header.append(cell.value)

		uid_index = header.index("Protein")
		pos_index = header.index("Position")

		total = ws.max_row

		for index, row in enumerate(ws):
			status("Reading input file...")
			if row != ws[1]:
				try:
					uid = row[uid_index].value[0:6]
				except TypeError: #this error occurs when "None" is present in uid cell
					continue #If you find "None" in the uid cell, skip everything below, and move on to the next row.
				if uid != '':
					numlines += 1
					pos = int(row[pos_index].value)
					if uid not in output_dict.keys():
						output_dict[uid] = [pos]
					else:
						output_dict[uid].append(pos)

			if progbar != None:
				#print(index)
				progbar(index/total*100)

	elif suffix == ".txt":
		pass
	else:
		print("Please supply a text, csv or excel file.")

	return output_dict, numlines


def update_databases():
	print("Nothing here")

def get_defaults():
	return [200, 20, 5] #[query length, max evalue, max pdbs reported]

def get_uniprot_seq(uid, resn, length, status=""):
	#status("This works.")
	#edge cases left to work out
	#What if the program isn't installed or can't be found?
	#What if the uniprot of interest isn't in the database?
	command_line = ['blastdbcmd', '-db', 'swissprot', '-dbtype', 'prot', '-entry', uid]
	#with open("/Users/acuesta/Box Sync/PDBSWS-test/stdout.txt", 'w+') as stdout:
		#with open("Users/acuesta/Box Sync/PDBSWS-test/stderr.txt", 'w+') as stderr:
	com = subprocess.run(command_line, capture_output = True, text=True)
	#com = subprocess.run(command_line, shell=True, stdout=stdout, stderr=stderr, stdin=subprocess.PIPE)
	results = com.stdout.split('\n')
	errors = com.stderr.split('\n')
	#print(results)
	for line in errors:
		if line != '':
			print(errors)
	seq = ""
	for line in results:
		if ">" not in line:
			seq += line.strip()


	if len(seq) <= length:
		return seq, seq
	elif resn > len(seq):
		print("Residue number is greater than sequence length.")
		return seq, seq
	elif resn < 1:
		print ("Residue number is zero or negative.")
		return seq, seq
	elif ((length/2 + resn) >= len(seq)):
		return seq[-length:], seq
	elif ((resn - length/2) <= 0):
		return seq[:length], seq
	else:
		return seq[int(resn-length/2):int(resn+length/2)], seq


	#print(seq)
	#print(results)

#returns a list of up to 5 PDB ids that meet a given e-value threshold
def pdb_uniprot_map(aa, uid, evalthreshold, percentthreshold):

	#submit query--limit to the the top 3 structures
	#get RID and get estimate
	#wait double the estimate
	#parse
	temp_fasta = open("query.fasta", 'w')
	temp_fasta.write(">%s\n%s" % (uid, aa))
	temp_fasta.close()
	command_line = ['blastp', '-db', 'pdbaa', '-query', "query.fasta", '-outfmt', "10 sacc qacc stitle evalue length bitscore nident pident"]
	com = subprocess.run(command_line, capture_output=True, text=True)
	results = com.stdout
	error = com.stderr
	#print(results)
	#print(error)

	output = []
	for line in results.split("\n"):
		output.append(line.split(","))
	#print(output)
	hs_pdbs = []
	eval_dict = {}
	percent_dict = {}

	for line in output:
		if len(line) > 1:
			#print(line)
			eval = line[-5]
			pident = float(line[-1])
			pdb = line[0].strip()
			if "-" in eval:
				#print(eval)
				eval_exp = int(eval[eval.index("-") + 1 :])
				#print(threshold)
				#print(eval_exp)
				if eval_exp >= evalthreshold and pident >= percentthreshold:
					#print("Here!")
					hs_pdbs.append(pdb)
					eval_dict[pdb] = eval
					percent_dict[pdb] = pident


	return hs_pdbs, eval_dict, percent_dict

def write_output(inputname, outputpath, filename, mapping, eval_dict, percent_dict, maxoutput=5, progbar=None, status=""):
	#print(outputname)
	outputname = outputpath + filename
	suffix = inputname.strip()[inputname.index("."):].lower()
	if suffix == ".csv":
		with open(inputname, 'r') as inputfile:
			reader = csv.reader(inputfile)
			header = next(reader)
			header.append("PDB IDs")
			with open(outputname, "w") as outputfile:
				with open("Hyperlinked_%s" % outputname, "w") as linkedfile:
					plainwriter = csv.writer(outputfile)
					linkedwriter = csv.writer(linkedfile)
					plainwriter.writerow(header)
					linkedwriter.writerow(header)

					uid_index = header.index("Protein")
					pos_index = header.index("Position")
					for row in reader:
						if len(row) != 0:
							#print(row)
							uid = row[uid_index][0:6]
							if uid != '':
								pos = int(row[pos_index])
								pdblist = mapping[uid][pos][:maxoutput]
								#print(pdblist)
								plainwriter.writerow(row + [pdblist])
								linked =[]
								for each in pdblist:
									linked.append("=Hyperlink(\"https://rcsb.org/structure/%s\", \"%s\")" % (each[0:4], each[0:4]))

								if len(linked) > 0:
									linkedwriter.writerow(row +linked)
								else:
									linkedwriter.writerow(row + ["None found"])




	elif suffix == ".xlsx" or suffix ==".xls":
		#print("writing")
		if status != "":
			status("Writing output...")
		wb_input = load_workbook(inputname)
		ws_input = wb_input.active

		wb_output = Workbook()
		ws_output = wb_output.active

		header = []
		for cell in ws_input[1]:
			header.append(cell.value)

		ws_output.append(header + ["PDB IDs", "e-value", "Percent Identity"])

		uid_index = header.index("Protein")
		pos_index = header.index("Position")
		#ws_input.cell(row=1, column=len(header)+1).value="PDB IDs"


		for index, row in enumerate(ws_input):
			#print(len(row))
			if row != ws_input[1]:
				try:
					uid = row[uid_index].value[0:6]
				except TypeError: #this error occurs when "None" is present in uid cell
					continue #If you find "None" in the uid cell, skip everything below, and move on to the next row.
				if uid != '':
					pos = int(row[pos_index].value)
					pdblist = mapping[uid][pos][:maxoutput]
					for count, pdbid in enumerate(pdblist):
						#ws_input.cell(row=index+1, column=len(row)+count).value="=Hyperlink(\"https://rcsb.org/structure/%s\", \"%s\")" % (pdbid[0:4], pdbid[0:4])
						output_row = []
						for each in row:
							output_row.append(each.value)
						try:
							ws_output.append(output_row + ["=Hyperlink(\"https://rcsb.org/structure/%s\", \"%s\")" % (pdbid[0:4], pdbid[0:4]), eval_dict[uid][pos][pdbid], percent_dict[uid][pos][pdbid]])
						except TypeError as e:
							print(e)
							print(pdbid)
							print(eval_dict)
							print(percent_dict)

					if len(pdblist) == 0:
						#ws_input.cell(row=index+1, column=len(row)).value="None found"]
						output_row = []
						for each in row:
							output_row.append(each.value)
						ws_output.append(output_row + ["None Found", "None Found", "None Found"])
		if "." not in outputname:
			outputname += ".xlsx"

		wb_output.save(filename = outputname)
		if status != "":
			status("Done! Starting ligand search...")


	elif suffix == ".txt":
		pass
	else:
		print("Please supply a text, csv or excel file.")



def process(inputpath, outputpath, filename, settings_dict, progbar=None, status=""):
	pairs, numlines = parse_input(inputpath, progbar, status)
	pdb_map_dict = {}
	query_sequences = {}
	full_uniprot_sequences = {}
	evalue_dict = {}
	percent_dict = {}

	length = int(settings_dict["query_length"])
	eval = int(settings_dict["eval_cutoff"])
	percent = int(settings_dict["min_percent"])
	max = int(settings_dict["max_pdb"])


	total = 0
	for index, id in enumerate(pairs.keys()):
		for count, pos in enumerate(pairs[id]):
			if status != "":
				total +=1
				status("Running Blastp search..." )
			queryseq, fullseq = get_uniprot_seq(id, pos, length, status)
			#status("Got uniprot seq")
			try:
				query_sequences[id][pos] = queryseq
				full_uniprot_sequences[id][pos] = fullseq
			except KeyError:
				query_sequences[id] = {pos : queryseq}
				full_uniprot_sequences[id] = {pos: fullseq}

			matched_pdbs, evalues, percents = pdb_uniprot_map(queryseq, id, eval, percent)
			try:
				evalue_dict[id][pos] = evalues
				percent_dict[id][pos] = percents
			except:
				evalue_dict[id] = {pos : evalues}
				percent_dict[id] = {pos : percents}
			#status("Matching pdbs")
			#print(matched_pdbs)
			if id not in pdb_map_dict.keys():
				pdb_map_dict[id] = {pos:matched_pdbs}
			elif pos not in pdb_map_dict[id].keys():
				pdb_map_dict[id][pos] = matched_pdbs
			else:
				pdb_map_dict[id][pos].append(matched_pdbs)

			if progbar != None:
				#print(index)
				progbar((total+3)/numlines*100)

	write_output(inputpath, outputpath, filename, pdb_map_dict, evalue_dict, percent_dict, max, progbar, status)

	if "distance_cutoff" in settings_dict.keys():
		threading.Thread(target=lnr.find_ligands(pdb_map_dict, evalue_dict, percent_dict, full_uniprot_sequences, inputpath, settings_dict, outputpath, progbar, status)).start()

def main():
	import sys, os, subprocess, csv
	from openpyxl import load_workbook

	inputname = sys.argv[1]
	outputname = sys.argv[2]
	length = sys.argv[3]
	eval = sys.argv[4]
	max = sys.argv[5]

	settings = {"query_length" : length, "eval_cutoff" : eval, "max_pdb" : max, "min_percent" : percent}


	process(inputname, outputname, settings)
